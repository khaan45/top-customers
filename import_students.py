"""
Imports the "Students" sheet of Student_Database.xlsx into the `students`
table that /api/students/lookup queries against.

Expected columns in the sheet (as produced by the earlier Excel build):
    STUDENT_ID | STUDENT_NAME | TOTAL_TRANSACTIONS | TOTAL_SPEND_SLSH
    | FIRST_TRANSACTION | LAST_TRANSACTION | STATUS | MOBILE

Only STUDENT_ID, STUDENT_NAME, MOBILE, and STATUS are imported — the
transaction/spend columns stay in the Excel file as reference, not in the
voting database. Safe to re-run: it upserts by student_id.

Usage:
    pip install openpyxl psycopg2-binary --break-system-packages
    python import_students.py Student_Database.xlsx \
        --db postgresql://user:pass@host:5432/dbname
"""

import argparse
import re
import sys

import openpyxl
import psycopg2


def normalize_phone(raw):
    if not raw:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if digits.startswith("0"):
        digits = "252" + digits[1:]
    elif not digits.startswith("252"):
        digits = "252" + digits
    return digits


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path")
    parser.add_argument("--db", required=True, help="Postgres connection string")
    parser.add_argument("--sheet", default="Students")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)
    ws = wb[args.sheet]

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        idx_id = header.index("STUDENT_ID")
        idx_name = header.index("STUDENT_NAME")
        idx_mobile = header.index("MOBILE")
        idx_status = header.index("STATUS")
    except ValueError as e:
        print(f"Expected column not found in sheet header {header}: {e}")
        return 1

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    imported, skipped_no_id, skipped_bad_phone = 0, 0, 0
    to_insert = []

    for row in rows:
        if not row or all(v is None for v in row):
            continue

        student_id = row[idx_id]
        name = row[idx_name]
        mobile_raw = row[idx_mobile]
        status = row[idx_status] or "Active"

        if not student_id or str(student_id).strip() == "":
            skipped_no_id += 1
            continue

        phone = normalize_phone(mobile_raw)
        if not phone or len(phone) < 9:
            skipped_bad_phone += 1
            continue

        to_insert.append((str(student_id).strip(), str(name).strip(), phone, str(status).strip()))

    conn = psycopg2.connect(args.db)
    cur = conn.cursor()
    for student_id, full_name, phone, status in to_insert:
        cur.execute(
            """
            INSERT INTO students (student_id, full_name, phone_number, status, has_voted)
            VALUES (%s, %s, %s, %s, FALSE)
            ON CONFLICT (student_id) DO UPDATE
              SET full_name = EXCLUDED.full_name,
                  phone_number = EXCLUDED.phone_number,
                  status = EXCLUDED.status
            """,
            (student_id, full_name, phone, status),
        )
        imported += 1
    conn.commit()
    cur.close()
    conn.close()

    print(f"Imported/updated: {imported}")
    print(f"Skipped (no Student ID): {skipped_no_id}")
    print(f"Skipped (unusable phone number): {skipped_bad_phone}")


if __name__ == "__main__":
    sys.exit(main())
